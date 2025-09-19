import streamlit as st
import numpy as np
import pandas as pd
from PIL import Image, ImageOps, ImageFilter, ImageDraw, ImageEnhance
import cv2
import fitz  # PyMuPDF for PDF handling
import io
from skimage.metrics import structural_similarity as ssim
import imutils
import easyocr
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import re
import datetime

st.set_page_config(
    layout="wide",
    page_title="Label Comparator",
    page_icon=r"icon/favicon.ico"  # Favicon
)

# --- Custom CSS ---
st.markdown("""
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Roboto:wght@300;400;700&display=swap');
        [data-testid=stBaseButton-secondary] {background:#f4a303;color:#023C59;}
        [data-testid=stFileUploaderDropzone] {
            background:#f7f7f7;border:2px dashed #4f7cae;border-radius:10px;
            padding:20px;color:#2d4059;
            font-weight:bold;font-size:18px;height:200px;
        }
        [data-testid=stWidgetLabel] {color:#f4a303;}
        .title-container {
            display:flex;justify-content:center;text-align:center;margin:auto;width:655px;
        }
        .comparison-container {
            background: rgba(244, 163, 3, 0.1);
            border: 2px solid #f4a303;
            border-radius: 10px;
            padding: 15px;
            margin: 20px 0;
        }
        .ocr-container {
            background: rgba(34, 197, 94, 0.1);
            border: 2px solid #22c55e;
            border-radius: 10px;
            padding: 15px;
            margin: 20px 0;
        }
        .metric-container {
            background: rgba(45, 64, 89, 0.8);
            border: 1px solid #4f7cae;
            border-radius: 8px;
            padding: 15px;
            margin: 10px 0;
        }
        [data-testid="stFileUploaderDropzoneInstructions"] small,
        [data-testid="stFileUploaderDropzoneInstructions"] span{
            color:#000000;
        }
    </style>
""", unsafe_allow_html=True)

# --- Header Section ---
header_html = """
    <div class="title-container">
        <span style="font-size:31px;font-weight:bold;color:#033c59;">Label</span>
        <span style="font-size:31px;font-weight:bold;color:#f4a303;margin-left:10px;">Comparator Tool</span>
    </div>
    <div style="text-align:center;margin-top:10px;color:#ccc;">
        <p>Compare images and PDFs </p>
    </div>
"""
with st.container(border=True):
    st.markdown(header_html, unsafe_allow_html=True)

# --- OCR Reader Initialization ---
@st.cache_resource
def init_ocr_reader():
    """Initialize EasyOCR reader"""
    try:
        return easyocr.Reader(['en'], gpu=False)
    except Exception as e:
        st.error(f"OCR initialization failed: {e}")
        return None

# Initialize OCR reader
ocr_reader = init_ocr_reader()

# --- PDF to Image Conversion ---
@st.cache_data
def pdf_to_image(pdf_file, page_num=0, dpi=200):
    """Convert PDF page to PIL Image"""
    try:
        pdf_bytes = pdf_file.read()
        pdf_file.seek(0)
        pdf_document = fitz.open(stream=pdf_bytes, filetype="pdf")
        first_page = pdf_document[0]
        zoom = dpi / 72
        mat = fitz.Matrix(zoom, zoom)
        pix = first_page.get_pixmap(matrix=mat, alpha=False)
        img_data = pix.tobytes("png")
        image = Image.open(io.BytesIO(img_data))
        if image.mode != 'RGB':
            image = image.convert('RGB')
        total_pages = len(pdf_document)
        pdf_document.close()
        return image, total_pages
    except Exception as e:
        st.error(f"Error converting PDF: {e}")
        return None, 0

# --- Image Preprocessing ---
def preprocess_image(image, resize_to=None, enhance_contrast=False):
    """Preprocess image for better comparison"""
    if image is None:
        return None
    img = image.copy()
    if resize_to:
        img = img.resize(resize_to, Image.Resampling.LANCZOS)
    if enhance_contrast:
        enhancer = ImageEnhance.Contrast(img)
        img = enhancer.enhance(1.5)
    return img

# --- OCR Image Preprocessing ---
def apply_ocr_preprocessing(img, invert=False, grayscale=False, binary=False, denoise=False):
    """Apply preprocessing filters for better OCR"""
    if img is None:
        return None

    img = img.copy()

    if invert:
        if img.mode == 'RGBA':
            img = img.convert('RGB')
        img = ImageOps.invert(img)

    if grayscale:
        img = img.convert('L')

    if binary:
        if img.mode != 'L':
            img = img.convert('L')
        img = img.point(lambda x: 255 if x > 128 else 0, 'L')

    if denoise:
        img = img.filter(ImageFilter.MedianFilter(3))

    return img

# --- OCR Text Extraction ---
def extract_text_with_ocr(img, confidence_threshold=0.5):
    """Extract text using EasyOCR"""
    if img is None or ocr_reader is None:
        return []

    # Convert PIL to numpy array for EasyOCR
    np_img = np.array(img)

    try:
        # Extract text with EasyOCR
        results = ocr_reader.readtext(np_img, detail=1)

        extracted_texts = []
        for (bbox, text, confidence) in results:
            if confidence >= confidence_threshold:
                # Clean the text
                cleaned_text = text.strip()
                if cleaned_text:  # Only add non-empty text
                    extracted_texts.append({
                        'text': cleaned_text,
                        'confidence': confidence,
                        'bbox': bbox
                    })

        return extracted_texts
    except Exception as e:
        st.error(f"OCR extraction error: {e}")
        return []

# --- Image Alignment ---
def align_images(imageA, imageB, max_features=500, good_match_percent=0.15):
    """Align two images using ORB feature matching"""
    try:
        grayA = cv2.cvtColor(np.array(imageA), cv2.COLOR_RGB2GRAY)
        grayB = cv2.cvtColor(np.array(imageB), cv2.COLOR_RGB2GRAY)

        orb = cv2.ORB_create(max_features)
        keypointsA, descriptorsA = orb.detectAndCompute(grayA, None)
        keypointsB, descriptorsB = orb.detectAndCompute(grayB, None)

        if descriptorsA is None or descriptorsB is None:
            return imageA, False

        matcher = cv2.DescriptorMatcher_create(cv2.DESCRIPTOR_MATCHER_BRUTEFORCE_HAMMING)
        matches = matcher.match(descriptorsA, descriptorsB)
        matches.sort(key=lambda x: x.distance, reverse=False)

        numGoodMatches = int(len(matches) * good_match_percent)
        matches = matches[:numGoodMatches]

        if len(matches) < 4:
            return imageA, False

        points1 = np.zeros((len(matches), 2), dtype=np.float32)
        points2 = np.zeros((len(matches), 2), dtype=np.float32)

        for i, match in enumerate(matches):
            points1[i, :] = keypointsA[match.queryIdx].pt
            points2[i, :] = keypointsB[match.trainIdx].pt

        h, mask = cv2.findHomography(points1, points2, cv2.RANSAC)

        if h is None:
            return imageA, False

        height, width = grayB.shape[:2]
        aligned = cv2.warpPerspective(np.array(imageA), h, (width, height))

        return Image.fromarray(aligned), True
    except Exception as e:
        return imageA, False

# --- Difference Detection ---
def find_differences(imageA, imageB, threshold=0.8, min_area=100):
    """Find differences between two images using SSIM"""
    try:
        if imageA.size != imageB.size:
            imageB = imageB.resize(imageA.size, Image.Resampling.LANCZOS)

        grayA = cv2.cvtColor(np.array(imageA), cv2.COLOR_RGB2GRAY)
        grayB = cv2.cvtColor(np.array(imageB), cv2.COLOR_RGB2GRAY)

        score, diff = ssim(grayA, grayB, full=True)
        diff = (diff * 255).astype("uint8")

        thresh = cv2.threshold(diff, 0, 255, cv2.THRESH_BINARY_INV | cv2.THRESH_OTSU)[1]

        cnts = cv2.findContours(thresh.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        cnts = imutils.grab_contours(cnts)

        filtered_cnts = [c for c in cnts if cv2.contourArea(c) > min_area]

        bounding_boxes = []
        for c in filtered_cnts:
            x, y, w, h = cv2.boundingRect(c)
            bounding_boxes.append((x, y, w, h))

        return {
            'ssim_score': score,
            'difference_image': diff,
            'threshold_image': thresh,
            'bounding_boxes': bounding_boxes,
            'total_differences': len(bounding_boxes)
        }
    except Exception as e:
        st.error(f"Error finding differences: {e}")
        return None

# --- Text Comparison ---
def compare_extracted_texts(base_texts, comp_texts):
    """Compare extracted texts between base and comparison documents"""
    differences = []
    
    # Create lists of just the text content
    base_text_list = [item['text'] for item in base_texts]
    comp_text_list = [item['text'] for item in comp_texts]
    
    # Find texts in base that are not in comparison
    missing_in_comp = [text for text in base_text_list if text not in comp_text_list]
    
    # Find texts in comparison that are not in base
    missing_in_base = [text for text in comp_text_list if text not in base_text_list]
    
    # Find matching texts
    matching_texts = [text for text in base_text_list if text in comp_text_list]
    
    return {
        'missing_in_comparison': missing_in_comp,
        'missing_in_base': missing_in_base,
        'matching_texts': matching_texts,
        'base_text_count': len(base_text_list),
        'comp_text_count': len(comp_text_list),
        'match_count': len(matching_texts),
        'missing_count': len(missing_in_comp) + len(missing_in_base)
    }

# --- Visualization ---
def draw_differences(image, bounding_boxes, color=(255, 0, 0), thickness=3):
    """Draw bounding boxes on image"""
    img_with_boxes = image.copy()
    draw = ImageDraw.Draw(img_with_boxes)

    for i, (x, y, w, h) in enumerate(bounding_boxes):
        draw.rectangle([x, y, x + w, y + h], outline=color, width=thickness)
        text = f"Diff {i+1}"
        draw.text((x, max(0, y-20)), text, fill=color)

    return img_with_boxes

def create_side_by_side_comparison(imageA, imageB, bounding_boxes, title="Comparison"):
    """Create side-by-side comparison with differences marked"""
    imgA_marked = draw_differences(imageA, bounding_boxes, color=(255, 0, 0))
    imgB_marked = draw_differences(imageB, bounding_boxes, color=(255, 0, 0))
    return imgA_marked, imgB_marked

# Replace the create_ocr_excel_export function in your existing code with this updated version

def create_ocr_excel_export(base_data, comparison_data, text_comparisons, filename="ocr_analysis.xlsx"):
    """Create Excel with specific pattern: Row | Base | Child1 | Empty | Base | Child2 | Empty
    and highlight missing values in red with updated Missing Text Details format"""
    wb = Workbook()
    ws = wb.active
    ws.title = "OCR Analysis"

    # Define styles
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green for base
    blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")   # Light blue for comparison
    red_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")    # Light red for differences
    yellow_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid") # Light yellow for partial matches
    
    # Red font for missing text
    red_font = Font(color="FF0000", bold=True)  # Red color for text not found in base
    normal_font = Font(color="000000")  # Normal black color

    # Create headers with the specified pattern: Row | Base | Child1 | Empty | Base | Child2 | Empty
    headers = ["Row"]

    for i, comp_data in enumerate(comparison_data):
        headers.append(f"Base: {base_data['filename'] if base_data else 'N/A'}")
        headers.append(f"Child {i+1}: {comp_data['filename']}")
        headers.append("")  # Empty column

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        if "Base:" in str(header):
            cell.fill = green_fill
        elif "Child" in str(header):
            cell.fill = blue_fill

    # Extract base text values for comparison (normalize for case-insensitive comparison)
    base_texts_set = set()
    if base_data and base_data['ocr_results']:
        base_texts_set = {item['text'].strip().lower() for item in base_data['ocr_results'] if item['text'].strip()}

    # Get maximum number of text entries
    max_rows = 0
    if base_data and base_data['ocr_results']:
        max_rows = len(base_data['ocr_results'])

    for comp_data in comparison_data:
        if comp_data['ocr_results']:
            max_rows = max(max_rows, len(comp_data['ocr_results']))

    # Fill data with the specified pattern
    for row in range(max_rows):
        excel_row = row + 2  # Start from row 2 (after header)
        col = 1

        # Row number
        ws.cell(row=excel_row, column=col, value=row + 1)
        col += 1

        # For each comparison: Base | Child | Empty
        for i, comp_data in enumerate(comparison_data):
            # Base data
            base_text = ""
            base_confidence = ""
            if base_data and base_data['ocr_results'] and row < len(base_data['ocr_results']):
                ocr_item = base_data['ocr_results'][row]
                base_text = ocr_item['text']
                base_confidence = f"({ocr_item['confidence']:.2f})"

            # Combine text and confidence
            base_cell_value = f"{base_text} {base_confidence}" if base_confidence else base_text
            
            cell = ws.cell(row=excel_row, column=col, value=base_cell_value)
            cell.fill = green_fill
            cell.font = normal_font
            col += 1

            # Child data
            comp_text = ""
            comp_confidence = ""
            is_missing_in_base = False
            
            if comp_data['ocr_results'] and row < len(comp_data['ocr_results']):
                ocr_item = comp_data['ocr_results'][row]
                comp_text = ocr_item['text']
                comp_confidence = f"({ocr_item['confidence']:.2f})"
                
                # Check if this text exists in base (case-insensitive)
                if comp_text.strip().lower() not in base_texts_set and comp_text.strip():
                    is_missing_in_base = True

            # Combine text and confidence
            comp_cell_value = f"{comp_text} {comp_confidence}" if comp_confidence else comp_text
            
            cell = ws.cell(row=excel_row, column=col, value=comp_cell_value)
            cell.fill = blue_fill
            
            # Apply red font if text is missing in base
            if is_missing_in_base:
                cell.font = red_font
            else:
                cell.font = normal_font
            col += 1

            # Empty column
            col += 1

    # Add a summary sheet with text comparison results
    if text_comparisons:
        ws_summary = wb.create_sheet("Text Comparison Summary")
        
        # Summary headers
        summary_headers = ["Comparison", "Base Text Count", "Child Text Count", 
                          "Matching Texts", "Texts Missing in Child", "Texts Missing in Base", 
                          "Match Percentage"]
        
        for col, header in enumerate(summary_headers, 1):
            cell = ws_summary.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        # Add comparison results
        for i, comp_result in enumerate(text_comparisons):
            ws_summary.cell(row=i+2, column=1, value=f"Comparison {i+1}")
            ws_summary.cell(row=i+2, column=2, value=comp_result['base_text_count'])
            ws_summary.cell(row=i+2, column=3, value=comp_result['comp_text_count'])
            ws_summary.cell(row=i+2, column=4, value=comp_result['match_count'])
            ws_summary.cell(row=i+2, column=5, value=len(comp_result['missing_in_comparison']))
            ws_summary.cell(row=i+2, column=6, value=len(comp_result['missing_in_base']))
            
            # Calculate match percentage
            if comp_result['base_text_count'] > 0:
                match_percentage = (comp_result['match_count'] / comp_result['base_text_count']) * 100
            else:
                match_percentage = 0
            ws_summary.cell(row=i+2, column=7, value=f"{match_percentage:.2f}%")

        # Add detailed missing texts section with NEW FORMAT (matching your image)
        ws_details = wb.create_sheet("Missing Text Details")
        
        # New headers based on your image format
        details_headers = ["Comparison", "Texts Missing in variant.jpg (present in Base.jpg)", "Texts Missing in Base.jpg (present in variant.jpg)"]
        
        for col, header in enumerate(details_headers, 1):
            cell = ws_details.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        # Add detailed missing texts - NEW FORMAT: One row per missing text
        detail_row = 2
        for i, comp_result in enumerate(text_comparisons):
            comp_name = f"{base_data['filename'] if base_data else 'Base'} vs {comparison_data[i]['filename']}"
            
            # First, add all texts missing in comparison (present in base)
            if comp_result['missing_in_comparison']:
                for missing_text in comp_result['missing_in_comparison']:
                    ws_details.cell(row=detail_row, column=1, value=comp_name)
                    ws_details.cell(row=detail_row, column=2, value=missing_text)
                    ws_details.cell(row=detail_row, column=3, value="")  # Empty for this type
                    detail_row += 1
            
            # Then, add all texts missing in base (present in child) - these should be in red
            if comp_result['missing_in_base']:
                for missing_text in comp_result['missing_in_base']:
                    ws_details.cell(row=detail_row, column=1, value=comp_name)
                    ws_details.cell(row=detail_row, column=2, value="")  # Empty for this type
                    cell = ws_details.cell(row=detail_row, column=3, value=missing_text)
                    cell.font = red_font  # Highlight in red
                    detail_row += 1
            
            # If no missing texts, add a row indicating this
            if not comp_result['missing_in_comparison'] and not comp_result['missing_in_base']:
                ws_details.cell(row=detail_row, column=1, value=comp_name)
                ws_details.cell(row=detail_row, column=2, value="No missing texts")
                ws_details.cell(row=detail_row, column=3, value="No missing texts")
                detail_row += 1

    # Adjust column widths
    for sheet in wb.worksheets:
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 60)  # Increased max width for longer text
            sheet.column_dimensions[column_letter].width = adjusted_width

    # Save to bytes
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    return excel_buffer.getvalue()
# Replace the Excel export function and the export button section with this:

# Replace the Excel export function with this updated version that includes spacing:

def create_missing_text_details_only_excel(base_data, comparison_data, text_comparisons, filename="missing_text_details.xlsx"):
    """Create Excel with only Missing Text Details tab, red text formatting, and empty rows after each variant"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Missing Text Details"

    # Define styles
    red_font = Font(color="FF0000", bold=True)  # Red color for all text content
    normal_font = Font(color="000000")  # Normal black color for headers and column A
    header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

    # Headers based on your image format
    details_headers = ["Comparison", "Texts Missing in variant.jpg (present in Base.jpg)", "Texts Missing in Base.jpg (present in variant.jpg)"]
    
    # Write headers with normal black font
    for col, header in enumerate(details_headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="000000")  # Black font for headers
        cell.fill = header_fill
    
    # Add detailed missing texts - One row per missing text with RED formatting
    detail_row = 2
    for i, comp_result in enumerate(text_comparisons):
        comp_name = f"{base_data['filename'] if base_data else 'Base'} vs {comparison_data[i]['filename']}"
        
        # Track if we added any rows for this comparison
        added_rows_for_this_comparison = False
        
        # First, add all texts missing in comparison (present in base)
        if comp_result['missing_in_comparison']:
            for missing_text in comp_result['missing_in_comparison']:
                # Column A (Comparison) - Normal black font
                ws.cell(row=detail_row, column=1, value=comp_name).font = normal_font
                
                # Column B (Missing in variant) - RED font
                cell_b = ws.cell(row=detail_row, column=2, value=missing_text)
                cell_b.font = red_font
                
                # Column C - Empty
                ws.cell(row=detail_row, column=3, value="")
                detail_row += 1
                added_rows_for_this_comparison = True
        
        # Then, add all texts missing in base (present in child) - RED font
        if comp_result['missing_in_base']:
            for missing_text in comp_result['missing_in_base']:
                # Column A (Comparison) - Normal black font
                ws.cell(row=detail_row, column=1, value=comp_name).font = normal_font
                
                # Column B - Empty
                ws.cell(row=detail_row, column=2, value="")
                
                # Column C (Missing in Base) - RED font
                cell_c = ws.cell(row=detail_row, column=3, value=missing_text)
                cell_c.font = red_font
                detail_row += 1
                added_rows_for_this_comparison = True
        
        # If no missing texts, add a row indicating this with RED font
        if not comp_result['missing_in_comparison'] and not comp_result['missing_in_base']:
            # Column A (Comparison) - Normal black font
            ws.cell(row=detail_row, column=1, value=comp_name).font = normal_font
            
            # Column B - RED font
            cell_b = ws.cell(row=detail_row, column=2, value="No missing texts")
            cell_b.font = red_font
            
            # Column C - RED font
            cell_c = ws.cell(row=detail_row, column=3, value="No missing texts")
            cell_c.font = red_font
            detail_row += 1
            added_rows_for_this_comparison = True
        
        # Add empty row after each variant group (except for the last one)
        if added_rows_for_this_comparison and i < len(text_comparisons) - 1:
            # Add empty row for spacing
            ws.cell(row=detail_row, column=1, value="")
            ws.cell(row=detail_row, column=2, value="")
            ws.cell(row=detail_row, column=3, value="")
            detail_row += 1

    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 60)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to bytes
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    return excel_buffer.getvalue()
def image_to_bytes(image):
    """Convert PIL Image to bytes"""
    with io.BytesIO() as buf:
        image.save(buf, format="PNG")
        return buf.getvalue()

# --- Main Application ---
def main():
    ss = st.session_state

    # Initialize session state
    if 'base_image' not in ss:
        ss.base_image = None
    if 'base_filename' not in ss:
        ss.base_filename = None
    if 'comparison_images' not in ss:
        ss.comparison_images = []
    if 'results' not in ss:
        ss.results = []
    if 'ocr_results' not in ss:
        ss.ocr_results = {}
    if 'text_comparisons' not in ss:
        ss.text_comparisons = []

    # Check OCR availability
    if ocr_reader is None:
        st.warning("⚠️ EasyOCR not available. Install with: pip install easyocr")

    # Sidebar for settings
    with st.sidebar:
        st.image("icon/Logo-final.png", width=250)
        st.header("⚙️ Comparison Settings")

        # Comparison parameters
        st.subheader("Detection Parameters")
        threshold = st.slider("SSIM Threshold", 0.0, 1.0, 0.8, 0.01, 
                             help="Higher values = more sensitive to differences")
        min_area = st.slider("Minimum Difference Area", 10, 1000, 100, 10,
                           help="Minimum pixel area for difference detection")

        # Image preprocessing
        st.subheader("Image Processing")
        enhance_contrast = st.checkbox("Enhance Contrast", help="Improve difference detection")
        align_images_option = st.checkbox("Auto-align Images", value=True, 
                                        help="Automatically align images before comparison")

        # OCR settings
        if ocr_reader:
            st.subheader("OCR Settings")
            confidence_threshold = st.slider("OCR Confidence Threshold", 0.1, 1.0, 0.5, 0.05,
                                           help="Minimum confidence for text extraction")

            # OCR preprocessing options
            st.subheader("OCR Preprocessing")
            ocr_invert = st.checkbox("Invert Colors", help="White text on dark background")
            ocr_grayscale = st.checkbox("Convert to Grayscale", help="Remove color information")
            ocr_binary = st.checkbox("Binary Threshold", help="Black and white only")
            ocr_denoise = st.checkbox("Denoise Image", help="Remove noise artifacts")

        # Export options
        st.subheader("Export Options")
        include_difference_images = st.checkbox("Include Difference Images", value=True)
        include_ocr_analysis = st.checkbox("Include OCR Analysis", value=True)

    # Main content area
    col1, col2 = st.columns([1, 1])

    with col1:
        st.subheader("📎 Base Document")
        base_file = st.file_uploader(
            "Upload base image or PDF", 
            type=["jpg", "png", "jpeg", "pdf"],
            key="base_upload"
        )

        if base_file:
            with st.spinner("Processing base document..."):
                if base_file.type == "application/pdf":
                    base_img, total_pages = pdf_to_image(base_file)
                    if base_img:
                        st.success(f"✓ Loaded PDF (Page 1 of {total_pages})")
                        ss.base_image = base_img
                        ss.base_filename = f"{base_file.name} (Page 1)"
                else:
                    base_img = Image.open(base_file)
                    if base_img.mode != 'RGB':
                        base_img = base_img.convert('RGB')
                    ss.base_image = base_img
                    ss.base_filename = base_file.name
                    st.success(f"✓ Loaded image: {base_file.name}")

    with col2:
        st.subheader("📋 Comparison Documents")
        comparison_files = st.file_uploader(
            "Upload images or PDFs to compare", 
            type=["jpg", "png", "jpeg", "pdf"],
            accept_multiple_files=True,
            key="comparison_upload"
        )

        if comparison_files:
            ss.comparison_images = []
            for file in comparison_files:
                with st.spinner(f"Processing {file.name}..."):
                    if file.type == "application/pdf":
                        comp_img, total_pages = pdf_to_image(file)
                        if comp_img:
                            ss.comparison_images.append((comp_img, f"{file.name} (Page 1)"))
                            st.success(f"✓ Loaded PDF: {file.name}")
                    else:
                        comp_img = Image.open(file)
                        if comp_img.mode != 'RGB':
                            comp_img = comp_img.convert('RGB')
                        ss.comparison_images.append((comp_img, file.name))
                        st.success(f"✓ Loaded image: {file.name}")

    # Analysis buttons
    if ss.base_image and ss.comparison_images:
        st.markdown("---")

        run_full_analysis = st.button("🚀 Run Comparator Analysis", type="primary")

        if run_full_analysis:
            # Visual comparison
            ss.results = []
            progress_bar = st.progress(0)
            total_comparisons = len(ss.comparison_images)

            for i, (comp_img, filename) in enumerate(ss.comparison_images):
                with st.spinner(f"Comparing with {filename}..."):
                    base_processed = preprocess_image(ss.base_image, enhance_contrast=enhance_contrast)
                    comp_processed = preprocess_image(comp_img, enhance_contrast=enhance_contrast)

                    if align_images_option:
                        comp_aligned, aligned_success = align_images(base_processed, comp_processed)
                        if not aligned_success:
                            comp_aligned = comp_processed
                    else:
                        comp_aligned = comp_processed

                    diff_results = find_differences(base_processed, comp_aligned, threshold, min_area)

                    if diff_results:
                        base_marked, comp_marked = create_side_by_side_comparison(
                            base_processed, comp_aligned, diff_results['bounding_boxes']
                        )

                        ss.results.append({
                            'filename': filename,
                            'base_marked': base_marked,
                            'comp_marked': comp_marked,
                            'difference_image': diff_results['difference_image'],
                            'ssim_score': diff_results['ssim_score'],
                            'total_differences': diff_results['total_differences'],
                            'bounding_boxes': diff_results['bounding_boxes']
                        })

                progress_bar.progress((i + 1) / total_comparisons)

            st.success(f"✅ Visual comparison completed!")

        if (run_full_analysis) and ocr_reader:
            with st.container():
                progress_bar_ocr = st.progress(0)
                total_documents = 1 + len(ss.comparison_images)
                current_doc = 0

                # Extract text from base image
                st.info("🔍 Extracting text from base document...")
                base_processed_ocr = apply_ocr_preprocessing(
                    ss.base_image, ocr_invert, ocr_grayscale, ocr_binary, ocr_denoise
                )
                base_ocr_results = extract_text_with_ocr(base_processed_ocr, confidence_threshold)

                ss.ocr_results['base'] = {
                    'filename': ss.base_filename,
                    'ocr_results': base_ocr_results,
                    'processed_image': base_processed_ocr
                }

                current_doc += 1
                progress_bar_ocr.progress(current_doc / total_documents)

                # Extract text from comparison images
                ss.ocr_results['comparisons'] = []
                ss.text_comparisons = []

                for i, (comp_img, filename) in enumerate(ss.comparison_images):
                    st.info(f"🔍 Extracting text from {filename}...")

                    comp_processed_ocr = apply_ocr_preprocessing(
                        comp_img, ocr_invert, ocr_grayscale, ocr_binary, ocr_denoise
                    )
                    comp_ocr_results = extract_text_with_ocr(comp_processed_ocr, confidence_threshold)

                    comp_data = {
                        'filename': filename,
                        'ocr_results': comp_ocr_results,
                        'processed_image': comp_processed_ocr
                    }
                    
                    ss.ocr_results['comparisons'].append(comp_data)
                    
                    # Compare texts between base and this comparison
                    text_comparison = compare_extracted_texts(
                        base_ocr_results, comp_ocr_results
                    )
                    ss.text_comparisons.append(text_comparison)

                    current_doc += 1
                    progress_bar_ocr.progress(current_doc / total_documents)

                # st.success(f"✅ OCR analysis completed for {total_documents} documents!")

    # Display Visual Comparison Results
    if ss.results:
        st.markdown("---")
        st.header("📊 Visual Comparison Results")

        tab_titles = [result['filename'] for result in ss.results]
        tabs = st.tabs(tab_titles)

        for tab, result in zip(tabs, ss.results):
            with tab:
                st.markdown(f"### 📄 {result['filename']}")

                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("SSIM Score", f"{result['ssim_score']:.4f}")
                with col2:
                    st.metric("Differences Found", result['total_differences'])
                with col3:
                    similarity_percent = result['ssim_score'] * 100
                    st.metric("Similarity %", f"{similarity_percent:.2f}%")

                colA, colB = st.columns(2)

                with colA:
                    st.markdown("**Base Document (with differences marked)**")
                    st.image(result['base_marked'])
                    st.download_button(
                        "📥 Download Base with Differences",
                        image_to_bytes(result['base_marked']),
                        f"base_marked_{result['filename']}.png",
                        "image/png",
                        key=f"base_marked_{result['filename']}"
                    )

                with colB:
                    st.markdown(f"**{result['filename']} (with differences marked)**")
                    st.image(result['comp_marked'])
                    st.download_button(
                        "📥 Download Comparison with Differences",
                        image_to_bytes(result['comp_marked']),
                        f"comparison_marked_{result['filename']}.png",
                        "image/png",
                        key=f"comp_marked_{result['filename']}"
                    )

        # Export Options
        st.markdown("---")
        st.subheader("📦 Export Report")

        col1, col2 = st.columns(2)

        # with col1:
        #     if st.button("📊 Generate Excel Report with Highlighting"):
        #         try:
        #             excel_data = create_ocr_excel_export(
        #                 ss.ocr_results['base'],
        #                 ss.ocr_results.get('comparisons', []),
        #                 ss.text_comparisons
        #             )
        #             timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        #             filename = f"label_comparator_highlighted_{timestamp}.xlsx"
        #             st.download_button(
        #                 label="📥 Download Excel Report",
        #                 data=excel_data,
        #                 file_name=filename,
        #                 mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        #                 key="excel_download_highlighted"
        #             )
        #             st.success("✅ Excel report with highlighting generated!")
        #             # st.info("🔴 Child values not found in Base are highlighted in RED")
        #             # st.info("📋 Additional sheets: 'Text Comparison Summary' and 'Missing Text Details'")
        #         except Exception as e:
        #             st.error(f"Error generating Excel: {e}")
    

        with col1:
            # if st.button("📊 Generate Report"):
            try:
                excel_data = create_missing_text_details_only_excel(
                    ss.ocr_results['base'],
                    ss.ocr_results.get('comparisons', []),
                    ss.text_comparisons
                )
                timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"Label_comparator_{timestamp}.xlsx"
                st.download_button(
                    label="📥 Download Report",
                    data=excel_data,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="missing_text_download"
                )
            
            except Exception as e:
                st.error(f"Error generating Excel: {e}")
if __name__ == "__main__":
    main()
